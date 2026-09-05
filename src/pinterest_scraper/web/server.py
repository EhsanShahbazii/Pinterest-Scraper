"""Web UI backend: FastAPI app that drives the scraper as background jobs.

Endpoints:
    GET  /                       -> web UI (index.html)
    GET  /api/health             -> liveness probe
    POST /api/scrape             -> start a scrape job, returns {job_id}
    GET  /api/jobs/{id}/events   -> Server-Sent Events with live progress
    GET  /api/jobs/{id}/result   -> final pins JSON when the job finishes
    POST /api/jobs/{id}/cancel   -> request cancellation
    GET  /api/jobs/{id}/images/{name} -> serve a downloaded image file
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import time
import zipfile
import queue
import threading
import uuid
from pathlib import Path

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, ConfigDict, Field

from ..dedupe import DedupeStore
from ..downloader import download_all
from ..http import build_session
from ..scraper import (board_pins, enrich_with_details, related_pins,
                       search_pins, typeahead_suggestions)
from ..storage import save_outputs

STATIC_DIR = Path(__file__).parent / "static"


class ScrapeRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")

    mode: str = Field(default="search", pattern="^(search|board)?$")
    query: str = Field(min_length=1)
    limit: int = Field(default=25, ge=1, le=500)
    download: bool = True
    details: bool = True
    dedup: bool = False
    workers: int = Field(default=4, ge=1, le=16)
    delay: float = Field(default=1.0, ge=0, le=30)
    jitter: float = Field(default=0.5, ge=0, le=10)
    batch_size: int = Field(default=10, ge=1, le=100)
    min_width: int = Field(default=0, ge=0, le=10000)
    min_height: int = Field(default=0, ge=0, le=10000)
    proxy: str = ""


class Job:
    def __init__(self, req: ScrapeRequest, out_dir: Path):
        self.id = uuid.uuid4().hex[:12]
        self.req = req
        self.out_dir = out_dir
        self.out_dir.mkdir(parents=True, exist_ok=True)
        self.events: queue.Queue = queue.Queue()
        self.status = "queued"
        self.phase = ""
        self.pins: list[dict] = []
        self.stats: dict = {}
        self.error = ""
        self.cancelled = False
        self.done = threading.Event()

    def emit(self, **ev):
        self.events.put(json.dumps(ev, default=str, ensure_ascii=False))

    def cancel_requested(self):
        return self.cancelled

    def _check_cancel(self):
        if self.cancelled:
            raise RuntimeError("cancelled by user")

    def run(self):
        try:
            self._run()
            self.status = "cancelled" if self.cancelled else "finished"
        except Exception as e:  # noqa: BLE001
            self.status = "cancelled" if "cancelled" in str(e) else "error"
            if self.status == "error":
                self.error = str(e)
        self.done.set()
        self.emit(event="done", status=self.status, error=self.error,
                  total=len(self.pins), stats=self.stats)

    def _run(self):
        req = self.req
        session = build_session(
            proxy_pool=[p.strip() for p in req.proxy.split(",") if p.strip()] or None)
        store = (DedupeStore(self.out_dir / ".seen_pins.json", scan_dir=self.out_dir)
                 if req.dedup else None)
        queries = [q.strip() for q in req.query.split(",") if q.strip()] or [req.query]
        batch = len(queries) > 1
        stem = (queries[0] if not batch else queries[0] + "-batch")
        stem = stem.strip().replace(" ", "_")[:40] or "pins"

        def run_one(query: str) -> list[dict]:
            self.emit(event="query_start", query=query,
                      index=queries.index(query) + 1, total=len(queries))
            is_board = (req.mode == "board")
            if is_board:
                return board_pins(session, query, req.limit,
                                  delay=req.delay, save_cb=batch_save,
                                  batch_size=req.batch_size, jitter=req.jitter)
            return search_pins(session, query, req.limit,
                               delay=req.delay, save_cb=batch_save,
                               batch_size=req.batch_size, jitter=req.jitter)

        collect_total = req.limit * len(queries)
        self.emit(event="phase", phase="collect", total=collect_total,
                  message="Collecting pins")

        def batch_save(partial: list[dict]) -> None:
            self._check_cancel()
            new = [p for p in partial if p["pin_id"] not in
                   {q["pin_id"] for q in self.pins}]
            if store:
                new = store.filter(new)
            self.pins.extend(new)
            self.emit(event="progress", phase="collect", count=len(self.pins),
                      total=collect_total)
            save_outputs(self.pins, self.out_dir, stem)  # batch-safe on-disk save

        existing_ids = {p["pin_id"] for p in self.pins}
        raw_pins: list[dict] = []
        for q in queries:
            for p in run_one(q):
                if p["pin_id"] not in existing_ids:
                    existing_ids.add(p["pin_id"])
                    raw_pins.append(p)
            self._check_cancel()

        pins = list(raw_pins)

        if not pins:
            clean_stem = re.sub(r"[^\w\-]+", "_", stem.lower().strip())
            for candidate_name in (f"{stem}.json", f"{clean_stem}.json", f"{stem.lower()}.json"):
                cached_file = self.out_dir / candidate_name
                if cached_file.exists():
                    try:
                        loaded = json.loads(cached_file.read_text(encoding="utf-8"))
                        if isinstance(loaded, list) and loaded:
                            pins = loaded
                            break
                    except Exception:
                        pass

        img_dir = self.out_dir / "images"
        if img_dir.exists():
            for p in pins:
                if not p.get("local_file"):
                    for ext in (".jpg", ".png", ".webp", ".jpeg"):
                        candidate = img_dir / f"{p['pin_id']}{ext}"
                        if candidate.is_file() and candidate.stat().st_size > 0:
                            p["local_file"] = candidate.name
                            break

        self.pins = pins

        if not pins:
            self.emit(event="nothing_new", total=0)
            return

        if req.details:
            self.emit(event="phase", phase="details", total=len(pins))
            done = {"n": 0}

            def _detail_cb(n):
                if self.cancelled:
                    raise RuntimeError("cancelled by user")
                done["n"] = max(done["n"], n)
                self.emit(event="progress", phase="details",
                          count=done["n"], total=len(pins))

            enrich_with_details(session, pins, delay=req.delay,
                                workers=req.workers, jitter=req.jitter,
                                progress_cb=_detail_cb)
            self._check_cancel()

        if req.download:
            self.emit(event="phase", phase="download", total=len(pins))
            from ..downloader import download_all

            def _dl_cb(n):
                if self.cancelled:
                    raise RuntimeError("cancelled by user")
                self.emit(event="progress", phase="download",
                          count=n, total=len(pins))

            self.stats = download_all(session, pins, self.out_dir / "images",
                                      req.workers, req.min_width, req.min_height,
                                      progress_cb=_dl_cb)
            self._check_cancel()

        summary = save_outputs(pins, self.out_dir, stem)
        if store:
            store.add(pins)
            store.save()
        self.pins = pins
        self.emit(event="saved", json_file=summary.get("json", ""),
                  csv_file=summary.get("csv", ""))


JOBS: dict[str, Job] = {}


def _job_out_dir() -> Path:
    return Path("web_output")


app = FastAPI(title="Pinterest Scraper Web")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "index.html")


@app.get("/api/health")
def health():
    return {"ok": True}


@app.post("/api/scrape")
def start_scrape(req: ScrapeRequest):
    job = Job(req, _job_out_dir())
    JOBS[job.id] = job
    threading.Thread(target=job.run, daemon=True).start()
    return {"job_id": job.id}


@app.get("/api/jobs/{job_id}/events")
async def job_events(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "unknown job")

    async def gen():
        loop = asyncio.get_running_loop()
        while True:
            try:
                ev = await asyncio.wait_for(
                    loop.run_in_executor(None, job.events.get, True, 0.2), 5)
                yield f"data: {ev}\n\n"
                if json.loads(ev).get("event") == "done":
                    return
            except (queue.Empty, asyncio.TimeoutError):
                if job.done.is_set() and job.events.empty():
                    return
                yield ": keepalive\n\n"

    return StreamingResponse(gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache"})


@app.get("/api/jobs/{job_id}/result")
def job_result(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "unknown job")
    if not job.done.is_set():
        raise HTTPException(409, "job still running")
    return {"status": job.status, "stats": job.stats, "error": job.error,
            "pins": job.pins}


@app.post("/api/jobs/{job_id}/cancel")
def cancel_job(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "unknown job")
    job.cancelled = True
    job.status = "cancelled"
    job.done.set()
    job.emit(event="done", status="cancelled", total=len(job.pins), stats=job.stats)
    return {"ok": True}


@app.get("/api/jobs/{job_id}/images/{name}")
def job_image(job_id: str, name: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "unknown job")
    path = (job.out_dir / "images" / name).resolve()
    if not str(path).startswith(str((job.out_dir).resolve())) or not path.is_file():
        raise HTTPException(404, "not found")
    return FileResponse(path)



# --- live search suggestions (cached, coalesced) ---
_SUGGEST_TTL = 300  # seconds
_suggest_cache: dict[str, tuple[float, list[str]]] = {}
_suggest_locks: dict[str, threading.Lock] = {}
_suggest_locks_guard = threading.Lock()


@app.get("/api/suggest")
def suggest(q: str = ""):
    q = q.strip().lower()
    if len(q) < 2:
        return {"suggestions": []}
    now = time.monotonic()
    hit = _suggest_cache.get(q)
    if hit and now - hit[0] < _SUGGEST_TTL:
        return {"suggestions": hit[1]}
    with _suggest_locks_guard:
        lock = _suggest_locks.setdefault(q, threading.Lock())
    with lock:  # coalesce concurrent identical requests
        now = time.monotonic()
        hit = _suggest_cache.get(q)
        if hit and now - hit[0] < _SUGGEST_TTL:
            return {"suggestions": hit[1]}
        try:
            session = build_session()
            out = typeahead_suggestions(session, q)
        except Exception:  # noqa: BLE001 — suggestions must never break the UI
            out = []
        _suggest_cache[q] = (now, out)
        if len(_suggest_cache) > 200:  # simple size cap
            oldest = sorted(_suggest_cache, key=lambda k: _suggest_cache[k][0])
            for k in oldest[:100]:
                _suggest_cache.pop(k, None)
        return {"suggestions": out}



@app.get("/api/images/{name}")
def global_image(name: str):
    img_dir = (_job_out_dir() / "images").resolve()
    path = (img_dir / name).resolve()
    if not path.is_relative_to(img_dir) or not path.is_file():
        raise HTTPException(404, "not found")
    return FileResponse(path)


@app.get("/api/gallery")
def get_gallery():
    img_dir = (_job_out_dir() / "images").resolve()
    if not img_dir.exists():
        return {"pins": [], "total": 0}

    metadata_map: dict[str, dict] = {}
    for jf in _job_out_dir().glob("*.json"):
        if jf.name == "schedules.json" or jf.name.startswith("."):
            continue
        try:
            items = json.loads(jf.read_text(encoding="utf-8"))
            if isinstance(items, list):
                for p in items:
                    if isinstance(p, dict) and p.get("pin_id"):
                        metadata_map[str(p["pin_id"])] = p
        except Exception:
            pass

    gallery_pins = []
    valid_exts = {".jpg", ".jpeg", ".png", ".webp"}
    image_files = sorted(
        [f for f in img_dir.iterdir() if f.is_file() and f.suffix.lower() in valid_exts],
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    for f in image_files:
        pin_id = f.stem
        meta = metadata_map.get(pin_id)
        if meta:
            p = dict(meta)
            p["local_file"] = f.name
        else:
            p = {
                "pin_id": pin_id,
                "title": f"Pin {pin_id}",
                "description": "",
                "local_file": f.name,
                "image_url": f"/api/images/{f.name}",
                "pin_url": f"https://www.pinterest.com/pin/{pin_id}/" if pin_id.isdigit() else "",
                "saves": None,
                "comments": None,
            }
        gallery_pins.append(p)

    return {"pins": gallery_pins, "total": len(gallery_pins)}


@app.get("/api/gallery/export/zip")
def export_gallery():
    img_dir = (_job_out_dir() / "images").resolve()
    valid_exts = {".jpg", ".jpeg", ".png", ".webp"}
    if not img_dir.exists():
        pins = []
    else:
        pins = [{"local_file": f.name} for f in img_dir.iterdir() if f.is_file() and f.suffix.lower() in valid_exts]
    return _zip_response(pins)


def _zip_response(pins: list[dict]) -> StreamingResponse:
    buf = io.BytesIO()
    img_dir = _job_out_dir() / "images"
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zf:
        for p in pins:
            f = p.get("local_file")
            if f and (img_dir / f).is_file():
                zf.write(img_dir / f, f"images/{f}")
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition":
                                      'attachment; filename="pins-images.zip"'})


def _xlsx_response(pins: list[dict]) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl is required for XLSX export. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "pins"
    cols = list(pins[0].keys()) if pins else ["pin_id"]
    ws.append(cols)
    for p in pins:
        ws.append([str(p.get(c)) if p.get(c) is not None else "" for c in cols])
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="pins.xlsx"'})


@app.get("/api/jobs/{job_id}/export/{fmt}")
def export_job(job_id: str, fmt: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "unknown job")
    if not job.done.is_set():
        raise HTTPException(409, "job still running")
    if fmt == "zip":
        return _zip_response(job.pins)
    if fmt == "xlsx":
        return _xlsx_response(job.pins)
    raise HTTPException(400, "format must be zip or xlsx")




# ---------------- visual search ----------------
_visual_cache: dict[str, tuple[float, list[dict]]] = {}


@app.get("/api/visual-search")
def visual_search(pin_id: str = "", limit: int = 25):
    pin_id = pin_id.strip()
    if not pin_id.isdigit():
        raise HTTPException(400, "numeric pin_id required")
    cached = _visual_cache.get(pin_id)
    if cached and time.monotonic() - cached[0] < 600:
        return {"pins": cached[1]}
    try:
        session = build_session()
        pins = related_pins(session, pin_id, limit=min(limit, 50))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"visual search failed: {e}")
    _visual_cache[pin_id] = (time.monotonic(), pins)
    return {"pins": pins}


# ---------------- scheduled scrapes ----------------
SCHEDULES_FILE = _job_out_dir() / "schedules.json"


def _load_schedules() -> list[dict]:
    if SCHEDULES_FILE.exists():
        try:
            return json.loads(SCHEDULES_FILE.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            return []
    return []


def _save_schedules(items: list[dict]) -> None:
    SCHEDULES_FILE.parent.mkdir(parents=True, exist_ok=True)
    SCHEDULES_FILE.write_text(json.dumps(items, ensure_ascii=False, indent=1), encoding="utf-8")


class ScheduleIn(BaseModel):
    mode: str = Field(default="search", pattern="^(search|board)$")
    query: str = Field(min_length=1)
    interval_hours: float = Field(default=24, ge=1, le=720)
    limit: int = Field(default=25, ge=1, le=200)


@app.get("/api/schedules")
def list_schedules():
    return {"schedules": _load_schedules()}


@app.post("/api/schedules")
def add_schedule(sch: ScheduleIn):
    items = _load_schedules()
    entry = {
        "id": uuid.uuid4().hex[:10],
        "mode": sch.mode, "query": sch.query.strip(),
        "interval_hours": sch.interval_hours, "limit": sch.limit,
        "next_run": time.time() + sch.interval_hours * 3600,
        "created": time.time(), "last_run": None, "runs": 0,
    }
    items.append(entry)
    _save_schedules(items)
    return entry


@app.delete("/api/schedules/{sid}")
def delete_schedule(sid: str):
    items = [s for s in _load_schedules() if s.get("id") != sid]
    _save_schedules(items)
    return {"ok": True, "remaining": len(items)}


def _scheduler_loop():
    """Background thread: run due schedules as normal dedup-on jobs."""
    while True:
        try:
            now = time.time()
            changed = False
            schedules = _load_schedules()
            for sch in schedules:
                if now >= sch.get("next_run", now + 3600):
                    req = ScrapeRequest(
                        mode=sch["mode"], query=sch["query"], limit=sch["limit"],
                        download=True, details=True, dedup=True)
                    job = Job(req, _job_out_dir())
                    JOBS[job.id] = job
                    threading.Thread(target=job.run, daemon=True).start()
                    sch["last_run"] = now
                    sch["next_run"] = now + sch["interval_hours"] * 3600
                    sch["runs"] = sch.get("runs", 0) + 1
                    sch["last_job_id"] = job.id
                    changed = True
            if changed:
                _save_schedules(schedules)
        except Exception:  # noqa: BLE001 — scheduler must never die
            pass
        time.sleep(60)


@app.on_event("startup")
def _start_scheduler():
    threading.Thread(target=_scheduler_loop, daemon=True).start()



# ---------------- image management ----------------
class DeleteImagesIn(BaseModel):
    names: list[str] = Field(default_factory=list)
    all: bool = False


@app.post("/api/images/delete")
def delete_images(payload: DeleteImagesIn):
    img_dir = (_job_out_dir() / "images").resolve()
    if payload.all:
        targets = [p for p in img_dir.glob("*") if p.is_file()]
    else:
        targets = []
        for name in payload.names:
            p = (img_dir / name).resolve()
            if str(p).startswith(str(img_dir)) and p.is_file():
                targets.append(p)
    deleted = 0
    for p in targets:
        try:
            p.unlink()
            deleted += 1
        except OSError:
            pass
    return {"deleted": deleted}




def main():
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000, log_level="warning")


if __name__ == "__main__":
    main()
